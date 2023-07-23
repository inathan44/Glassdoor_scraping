from openpyxl import Workbook, load_workbook
from datetime import datetime


def write_to_excel(
    file_name,
    company,
    title,
    url,
    size,
    industry,
    revenue,
    age,
    keyword_snippets,
    sheet_name,
):
    try:
        # Load existing workbook if it exists, or create a new one
        try:
            workbook = load_workbook(file_name)
        except FileNotFoundError:
            workbook = Workbook()

        date_value = datetime.now()
        formatted_date = date_value.strftime("%m-%d")

        sheet_name = sheet_name + "_" + formatted_date

        if sheet_name in workbook.sheetnames:  # Find or create sheet
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Add column headers if not already present
        if sheet.max_row == 1:
            column_headers = [
                "company",
                "title",
                "url",
                "size",
                "industry",
                "revenue",
                "age",
            ]

            # Add a column for each keyword
            for keyword in keyword_snippets.keys():
                column_headers.append(f"keyword: {keyword}")

            sheet.append(column_headers)

        # Append data to the end of the columns
        data = [company, title, url, size, industry, revenue, age]
        for keyword in keyword_snippets.keys():
            data.append(keyword_snippets[keyword])

        sheet.append(data)

        # Save the workbook
        workbook.save(file_name)
        print("Data written to Excel successfully.")

    except Exception as e:
        print("An error occurred while writing to Excel:", str(e))
